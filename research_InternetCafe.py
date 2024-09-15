import openpyxl
from openpyxl.styles import Border,Side,Font,PatternFill
import os,platform,sys
import numpy as np
import Dekutools

PYTHONVER_TUPLE=platform.python_version_tuple()

#数字が370
if not((int(PYTHONVER_TUPLE[0]) >=3 and int(PYTHONVER_TUPLE[1]) >= 7)or int(PYTHONVER_TUPLE[0])>=4):
    sys.stderr('[Error Code:-1]このスクリプトはPythonVer3.7以降でないと対応していません。')
    sys.exit(-1)

#Tは１０の倍数。
T=range(0,60*24+1,10)

FILEPATH=os.path.expanduser("~/Library/CloudStorage/OneDrive-個人用/Documents/ネットカフェ料金表.xlsm")
wb=openpyxl.load_workbook(FILEPATH,keep_vba=True)#expanduserは~をホームディレクトリに置き換える関数。
I_RANGE,O_RANGE=[wb['入力'],wb['レポート']]
O_RANGE.delete_rows(200)#初期化

A,C,L,L_Titles=[[] for val in range(4)]

for rowNum,row in enumerate(I_RANGE['A1:Z11']):
    l=[]
    isPackSec=rowNum>=3
    for colNum,col in enumerate(row):
        if colNum==0:
            continue
        elif col.value!=None:
            if rowNum==0:
                L_Titles.append(col.value)
            elif rowNum==1:
                C.append(col.value)
            elif rowNum==2:
                A.append(col.value)
            elif isPackSec:
                l.append(col.value) 
    if isPackSec:
        L.append(l)
#ブース、飲み放題カフェそれぞれのパックの料金を取り出す。
L=np.array(L).T

price=0
cnt=0
DicData={'時間(分)':list(T)}
#シートブース。
#名前をつける。
for KIND,l_title in zip(range(len(L)),L_Titles):
    prices=[]
    for t in T:

        if L is None:
            add_f=A[KIND]*(t-30)/10
            add_f=add_f if add_f>0 else 0
            price=add_f+C[KIND]
        elif t<180:
            add_f=A[KIND]*(t-30)/10
            add_f=add_f if add_f>0 else 0
            price=add_f+C[KIND]
            price=price if L[KIND][0]>price else L[KIND][0]
        #【将来】ネットカフェの８時間パックの料金を考慮した場合。
        elif t>=180*cnt and t<=180*(cnt+1):
            price=A[KIND]*(t-180*cnt)/10+L[KIND][cnt-1]
            price=price if L[KIND][cnt]>=price else L[KIND][cnt]
        cnt=int(t/180) if t%180==0 else cnt
        prices.append(int(price))
    DicData[l_title]=prices

cell_pos=None
for i,d_item in enumerate(DicData.items()):
    cell_pos=i+1
    #網掛け、罫線の設定。
    h_fill=PatternFill(patternType='solid',fgColor='0000ff')
    theFont=Font(bold=True,color='ffffff')
    theSide=Side(color='000000',style='thin')
    theBorder=Border(left=theSide,right=theSide,top=theSide,bottom=theSide)

    header_cell=O_RANGE.cell(column=cell_pos,row=1)
    header_cell.value=d_item[0]
    header_cell.fill=h_fill
    header_cell.font=theFont
    header_cell.border=theBorder
    for j,val in enumerate(d_item[1]):
        row_pos=j+2
        val_cell=O_RANGE.cell(column=cell_pos,row=row_pos)
        val_cell.border=theBorder
        val_cell.value=val
crossTableOriginPos=cell_pos+2
LASTROW_POS=row_pos

#クロス表の作成
CROSS_HEADERS=['予算（円)','使用予定時間（分）']
CROSS_HEADERS.extend(list(DicData)[1:])
for i,c_header in enumerate(CROSS_HEADERS):
    O_RANGE.cell(column=crossTableOriginPos+i,row=1).value=c_header
    selected_cell=O_RANGE.cell(column=crossTableOriginPos+i,row=1)
    selected_cell.border=theBorder
    selected_cell=O_RANGE.cell(column=crossTableOriginPos+i,row=2)
    selected_cell.border=theBorder
    if i>=2:
        #末尾から順に表をマッチングする。
        excel_func={
            "budget":\
                f'text(XLOOKUP(${Dekutools.ConvertToA1(crossTableOriginPos)}$2,' +\
            f'{Dekutools.ConvertToA1(i)}2:{Dekutools.ConvertToA1(i)}{LASTROW_POS},'+\
            f'$A$2:$A${LASTROW_POS},,1,-1)/60/24,"hh:mm")',

            "price":\
                f'round(XLOOKUP(${Dekutools.ConvertToA1(crossTableOriginPos+1)}$2,'+\
                f'$A$2:$A${LASTROW_POS},'+\
            f'{Dekutools.ConvertToA1(i)}2:{Dekutools.ConvertToA1(i)}{LASTROW_POS},'+\
            f',1,-1)'+
            ',2)',
        }
        func_val=f'=if(${Dekutools.ConvertToA1(crossTableOriginPos+1)}$2<>"",{excel_func["price"]},{excel_func["budget"]})'
        O_RANGE.cell(column=crossTableOriginPos+i,row=2).value=func_val
        selected_cell=O_RANGE.cell(column=crossTableOriginPos+i,row=2)
        selected_cell.border=theBorder
#グラフの作成はVBAにお任せ。

wb.save(FILEPATH)